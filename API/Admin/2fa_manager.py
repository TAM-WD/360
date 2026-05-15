# -*- coding: utf-8 -*-
"""
Yandex 360 — менеджер двухфакторной аутентификации (2FA).
"""

import os
import sys
import csv
import json
import time
import logging
import getpass
import datetime
from typing import List, Tuple

try:
    import requests
except ImportError:
    print("Модуль 'requests' не установлен. Установите: pip install requests")
    sys.exit(1)

# openpyxl — опционально (только для .xlsx)
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


def progress(iterable, desc: str, total=None):
    """Возвращает итератор с прогресс-баром (tqdm если установлен, иначе ASCII)."""
    if HAS_TQDM:
        return tqdm(iterable, desc=desc, total=total, ncols=80, unit="шт")
    if total is None:
        try:
            total = len(iterable)
        except TypeError:
            total = 0

    def _gen():
        bar_w = 30
        for i, item in enumerate(iterable, 1):
            yield item
            if total:
                done = int(bar_w * i / total)
                bar = "#" * done + "-" * (bar_w - done)
                sys.stdout.write(f"\r  {desc}: [{bar}] {i}/{total}")
                sys.stdout.flush()
        if total:
            sys.stdout.write("\n")
            sys.stdout.flush()
    return _gen()


ID_PREFIX = "11300000"

API_BASE = "https://api360.yandex.net"
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

log_file = os.path.join(LOG_DIR, f"2fa_manager_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("2fa")


# ---------------- Маскированный ввод ----------------

def masked_input(prompt: str, mask_char: str = "*") -> str:
    """
    Ввод строки с подменой символов на mask_char в реальном времени.
    На Windows — через msvcrt; на других ОС — fallback на getpass с фиксированной маской.
    """
    sys.stdout.write(prompt)
    sys.stdout.flush()

    if os.name == "nt":
        import msvcrt
        buf = []
        while True:
            ch = msvcrt.getwch()
            if ch in ("\r", "\n"):
                sys.stdout.write("\n")
                sys.stdout.flush()
                break
            elif ch == "\x03":  # Ctrl+C
                raise KeyboardInterrupt
            elif ch == "\b":  # Backspace
                if buf:
                    buf.pop()
                    sys.stdout.write("\b \b")
                    sys.stdout.flush()
            elif ch == "\x00" or ch == "\xe0":
                # спец-клавиша — съесть следующий символ
                msvcrt.getwch()
            else:
                buf.append(ch)
                sys.stdout.write(mask_char)
                sys.stdout.flush()
        return "".join(buf)
    else:
        # Кросс-платформенный fallback (без живой маски)
        value = getpass.getpass("")
        print(mask_char * len(value))
        return value


def display_masked(value: str, mask_char: str = "*") -> str:
    return mask_char * len(value)


def is_allowed_user_id(uid) -> bool:
    return str(uid).startswith(ID_PREFIX)


# ---------------- HTTP вспомогательные ----------------

class ApiClient:
    def __init__(self, token: str, org_id: str):
        self.token = token
        self.org_id = org_id
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"OAuth {token}",
            "Content-Type": "application/json",
        })

    def _request(self, method: str, url: str, **kwargs) -> requests.Response:
        for attempt in range(3):
            try:
                resp = self.session.request(method, url, timeout=30, **kwargs)
                if resp.status_code == 429:
                    wait = 2 ** attempt
                    log.warning(f"429 Too Many Requests, ждём {wait}с")
                    time.sleep(wait)
                    continue
                return resp
            except requests.RequestException as e:
                log.error(f"Сетевая ошибка ({attempt+1}/3): {e}")
                time.sleep(2 ** attempt)
        raise RuntimeError(f"Не удалось выполнить запрос {method} {url}")

    def list_users(self) -> List[dict]:
        """Постранично собирает всех пользователей организации с прогресс-баром."""
        users = []
        page = 1
        per_page = 1000
        url = f"{API_BASE}/directory/v1/org/{self.org_id}/users"

        # Первый запрос — узнаём общее число страниц.
        resp = self._request("GET", url, params={"page": page, "perPage": per_page})
        if resp.status_code != 200:
            raise RuntimeError(f"GET users -> {resp.status_code}: {resp.text}")
        data = resp.json()
        pages = data.get("pages", 1) or 1
        users.extend(data.get("users", []))
        log.info(f"Получено пользователей: страница {page}/{pages}, на странице {len(data.get('users', []))}")

        if pages > 1:
            bar = None
            if HAS_TQDM:
                bar = tqdm(total=pages, desc="Загрузка пользователей", ncols=80, unit="стр")
                bar.update(1)
            for page in range(2, pages + 1):
                resp = self._request("GET", url, params={"page": page, "perPage": per_page})
                if resp.status_code != 200:
                    raise RuntimeError(f"GET users -> {resp.status_code}: {resp.text}")
                data = resp.json()
                chunk = data.get("users", [])
                users.extend(chunk)
                log.info(f"Получено пользователей: страница {page}/{pages}, на странице {len(chunk)}")
                if bar is not None:
                    bar.update(1)
                else:
                    sys.stdout.write(f"\r  Загрузка пользователей: {page}/{pages}")
                    sys.stdout.flush()
                if not chunk:
                    break
            if bar is not None:
                bar.close()
            else:
                sys.stdout.write("\n")
                sys.stdout.flush()
        return users

    def enable_domain_2fa(self, duration_sec: int) -> dict:
        url = f"{API_BASE}/security/v2/org/{self.org_id}/domain_2fa"
        body = {
            "duration": duration_sec,
            "logoutUsers": False,
            "validationMethod": "default",
            "scope": "per_user",
        }
        resp = self._request("POST", url, data=json.dumps(body))
        if resp.status_code not in (200, 201):
            raise RuntimeError(f"POST domain_2fa -> {resp.status_code}: {resp.text}")
        return resp.json() if resp.text else {}

    def set_user_2fa(self, user_id: str, enabled: bool) -> Tuple[bool, str]:
        url = f"{API_BASE}/directory/v1/org/{self.org_id}/users/{user_id}/domain_2fa"
        params = {"is2faEnabled": "true" if enabled else "false"}
        resp = self._request("PATCH", url, params=params)
        if resp.status_code in (200, 204):
            return True, ""
        return False, f"{resp.status_code}: {resp.text}"

    def find_user_by_email(self, email: str) -> str:
        """Находит userId по email/nickname. Возвращает id или пустую строку."""
        # Простейший вариант — пройтись по всем пользователям (кешируется снаружи).
        return ""


# ---------------- Чтение списка из файла ----------------

def read_identifiers_from_file(path: str) -> List[str]:
    ext = os.path.splitext(path)[1].lower()
    items: List[str] = []
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(f, dialect)
            for row in reader:
                for cell in row:
                    v = (cell or "").strip()
                    if v:
                        items.append(v)
    elif ext in (".xlsx", ".xlsm"):
        if not HAS_OPENPYXL:
            raise RuntimeError("Для чтения .xlsx требуется openpyxl: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    v = str(cell).strip()
                    if v:
                        items.append(v)
    else:
        raise RuntimeError(f"Неподдерживаемое расширение файла: {ext}")

    # уникализация с сохранением порядка
    seen = set()
    result = []
    for v in items:
        if v.lower() in seen:
            continue
        seen.add(v.lower())
        # Пропускаем строки-заголовки
        if v.lower() in ("email", "uid", "id", "user", "user_id", "почта", "ящик"):
            continue
        result.append(v)
    return result


def resolve_to_user_ids(identifiers: List[str], all_users: List[dict]) -> Tuple[List[str], List[str]]:
    """Преобразует список email/UID в список userId. Возвращает (ids, unresolved)."""
    by_email = {}
    by_nickname = {}
    by_id = set()
    for u in all_users:
        uid = str(u.get("id", ""))
        by_id.add(uid)
        email = (u.get("email") or "").lower()
        if email:
            by_email[email] = uid
        for alias in u.get("aliases", []) or []:
            by_email[str(alias).lower()] = uid
        nick = (u.get("nickname") or "").lower()
        if nick:
            by_nickname[nick] = uid
        for c in u.get("contacts", []) or []:
            if c.get("type") == "email":
                v = (c.get("value") or "").lower()
                if v:
                    by_email[v] = uid

    ids, unresolved = [], []
    for raw in identifiers:
        v = raw.strip()
        if not v:
            continue
        if v.isdigit() and v in by_id:
            ids.append(v)
            continue
        low = v.lower()
        if low in by_email:
            ids.append(by_email[low])
        elif low.split("@")[0] in by_nickname:
            ids.append(by_nickname[low.split("@")[0]])
        else:
            unresolved.append(raw)
    return ids, unresolved


# ---------------- Сценарии ----------------

def _new_report() -> dict:
    return {
        "users_total": 0,
        "users_allowed": 0,
        "users_skipped_prefix": 0,
        "domain_2fa_enabled": False,
        "users_2fa_enabled_ok": 0,
        "users_2fa_enabled_fail": 0,
        "users_2fa_disabled_ok": 0,
        "users_2fa_disabled_fail": 0,
        "disabled_skipped_prefix": 0,
        "unresolved": [],
        "errors": [],
    }


def _filter_allowed(users: list, report: dict) -> list:
    """Оставляет только пользователей, чей ID начинается на ID_PREFIX.
    Остальных считает в report['users_skipped_prefix']."""
    allowed = []
    for u in users:
        uid = str(u.get("id", ""))
        if is_allowed_user_id(uid):
            allowed.append(u)
        else:
            report["users_skipped_prefix"] += 1
            log.info(f"Пропущен (ID не начинается на {ID_PREFIX}): {uid}")
    report["users_allowed"] = len(allowed)
    return allowed


def scenario_full(client: ApiClient) -> dict:
    report = _new_report()

    # Шаг 1
    print("\n[1/4] Получаем список пользователей организации...")
    users = client.list_users()
    report["users_total"] = len(users)
    print(f"  → найдено пользователей: {len(users)}")

    target_users = _filter_allowed(users, report)
    print(f"  → к обработке (ID начинается на {ID_PREFIX}): {len(target_users)}")
    print(f"  → пропущено (другой префикс ID): {report['users_skipped_prefix']}")

    # Шаг 2
    print("\n[2/4] Включаем 2FA на уровне организации (per_user).")
    while True:
        raw = input("  Введите период отсрочки (в секундах, целое число): ").strip()
        try:
            duration = int(raw)
            if duration < 0:
                raise ValueError
            break
        except ValueError:
            print("  Введите целое неотрицательное число.")
    try:
        client.enable_domain_2fa(duration)
        report["domain_2fa_enabled"] = True
        print(f"  → domain_2fa включён (duration={duration}s, scope=per_user).")
    except Exception as e:
        msg = f"Не удалось включить domain_2fa: {e}"
        log.error(msg)
        report["errors"].append(msg)
        print(f"  ! {msg}")

    # Шаг 3
    print(f"\n[3/4] Включаем 2FA пользователям ({len(target_users)} шт.)...")
    for u in progress(target_users, desc="Включение 2FA", total=len(target_users)):
        uid = str(u.get("id"))
        ok, err = client.set_user_2fa(uid, True)
        if ok:
            report["users_2fa_enabled_ok"] += 1
        else:
            report["users_2fa_enabled_fail"] += 1
            log.error(f"User {uid} enable 2fa failed: {err}")
            report["errors"].append(f"enable {uid}: {err}")

    # Шаг 4-5
    print("\n[4/4] Отключение 2FA выбранным пользователям.")
    identifiers = prompt_identifiers()
    if identifiers:
        _process_disable_identifiers(client, users, identifiers, report)
    else:
        print("  → шаг пропущен.")

    return report


def parse_identifiers_line(text: str) -> List[str]:
    """Парсит строку с email/UID, разделёнными пробелом / запятой / точкой с запятой / переводом строки."""
    if not text:
        return []
    raw = text.replace(",", " ").replace(";", " ").replace("\t", " ").replace("\n", " ")
    parts = [p.strip().strip('"').strip("'") for p in raw.split(" ")]
    out, seen = [], set()
    for p in parts:
        if not p:
            continue
        low = p.lower()
        if low in seen:
            continue
        seen.add(low)
        out.append(p)
    return out


def prompt_identifiers() -> List[str]:
    """Спрашивает у пользователя источник списка для отключения 2FA.
    Возвращает список идентификаторов (email/UID) или пустой список, если шаг пропущен."""
    while True:
        print("  Откуда взять список пользователей для отключения 2FA?")
        print("    1) Файл .csv / .xlsx")
        print("    2) Ввести вручную (через пробел, запятую или точку с запятой)")
        print("    0) Пропустить шаг")
        ans = input("  Ваш выбор [1/2/0]: ").strip()
        if ans == "0" or ans == "":
            return []
        if ans == "1":
            path = input("  Укажите путь к файлу: ").strip().strip('"')
            if not os.path.isfile(path):
                print("  ! Файл не найден, попробуйте снова.")
                continue
            try:
                ids = read_identifiers_from_file(path)
                print(f"  → прочитано записей: {len(ids)}")
                return ids
            except Exception as e:
                log.exception("Ошибка чтения файла")
                print(f"  ! Ошибка чтения файла: {e}")
                continue
        if ans == "2":
            print("  Введите email/UID через пробел, запятую или ';'.")
            print("  Для многострочного ввода завершите пустой строкой.")
            lines = []
            first = input("  > ")
            if first.strip():
                lines.append(first)
                while True:
                    nxt = input("  > ")
                    if not nxt.strip():
                        break
                    lines.append(nxt)
            ids = parse_identifiers_line("\n".join(lines))
            if not ids:
                print("  ! Не распознано ни одного идентификатора.")
                continue
            print(f"  → распознано записей: {len(ids)}")
            return ids
        print("  Неизвестный пункт.")


def _process_disable_identifiers(client: "ApiClient", users: list, identifiers: List[str], report: dict) -> None:
    """Преобразует идентификаторы в userId, фильтрует по префиксу и отключает 2FA."""
    try:
        ids, unresolved = resolve_to_user_ids(identifiers, users)
        report["unresolved"] = unresolved
        if unresolved:
            print(f"  ! не удалось сопоставить {len(unresolved)} записей (см. лог)")
            for u_ in unresolved:
                log.warning(f"Не сопоставлено: {u_}")

        filtered_ids = []
        for uid in ids:
            if is_allowed_user_id(uid):
                filtered_ids.append(uid)
            else:
                report["disabled_skipped_prefix"] += 1
                log.info(f"Пропущен при отключении (ID не начинается на {ID_PREFIX}): {uid}")
        if report["disabled_skipped_prefix"]:
            print(f"  → пропущено по префиксу ID: {report['disabled_skipped_prefix']}")
        print(f"  → к отключению: {len(filtered_ids)}")

        for uid in progress(filtered_ids, desc="Отключение 2FA", total=len(filtered_ids)):
            ok, err = client.set_user_2fa(uid, False)
            if ok:
                report["users_2fa_disabled_ok"] += 1
            else:
                report["users_2fa_disabled_fail"] += 1
                log.error(f"User {uid} disable 2fa failed: {err}")
                report["errors"].append(f"disable {uid}: {err}")
    except Exception as e:
        msg = f"Ошибка обработки списка: {e}"
        log.exception(msg)
        report["errors"].append(msg)
        print(f"  ! {msg}")


def print_report(report: dict) -> None:
    print("\n" + "=" * 60)
    print("ОТЧЁТ О РАБОТЕ СКРИПТА")
    print("=" * 60)
    print(f"Всего пользователей в организации       : {report['users_total']}")
    print(f"Пользователей с ID, начинающимся на {ID_PREFIX}: {report['users_allowed']}")
    print(f"Пропущено пользователей (другой префикс) : {report['users_skipped_prefix']}")
    print(f"domain_2fa включён                      : {'да' if report['domain_2fa_enabled'] else 'нет'}")
    print(f"Включено 2FA  (успех / ошибка)          : {report['users_2fa_enabled_ok']} / {report['users_2fa_enabled_fail']}")
    print(f"Отключено 2FA (успех / ошибка)          : {report['users_2fa_disabled_ok']} / {report['users_2fa_disabled_fail']}")
    if report.get("disabled_skipped_prefix"):
        print(f"При отключении пропущено по префиксу    : {report['disabled_skipped_prefix']}")
    if report["unresolved"]:
        print(f"Не сопоставлено записей                 : {len(report['unresolved'])}")
    if report["errors"]:
        print(f"Ошибок зафиксировано                    : {len(report['errors'])}")
        print(f"  Подробности в логе: {log_file}")
    print("=" * 60 + "\n")


# ---------------- Меню ----------------

def main_menu(token: str, org_id: str) -> None:
    client = ApiClient(token, org_id)
    while True:
        print("\n========== Yandex 360 — менеджер 2FA ==========")
        print(f" Токен      : {display_masked(token)}")
        print(f" Org ID     : {org_id}")
        print("------------------------------------------------")
        print(" 1. Запустить полный сценарий (1→5)")
        print(" 2. Только: получить список пользователей")
        print(" 3. Только: включить domain_2fa (per_user)")
        print(" 4. Только: включить 2FA всем пользователям")
        print(" 5. Только: отключить 2FA по списку (файл или ручной ввод)")
        print(" 0. Выход")
        choice = input("Ваш выбор: ").strip()

        try:
            if choice == "1":
                rep = scenario_full(client)
                print_report(rep)
            elif choice == "2":
                rep = _new_report()
                users = client.list_users()
                rep["users_total"] = len(users)
                _filter_allowed(users, rep)
                print(f"Получено пользователей: {len(users)} (с префиксом {ID_PREFIX}: {rep['users_allowed']})")
                print_report(rep)
            elif choice == "3":
                raw = input("Период отсрочки (сек): ").strip()
                duration = int(raw)
                client.enable_domain_2fa(duration)
                print("domain_2fa включён.")
            elif choice == "4":
                rep = _new_report()
                users = client.list_users()
                rep["users_total"] = len(users)
                target_users = _filter_allowed(users, rep)
                print(f"К обработке: {len(target_users)} (пропущено {rep['users_skipped_prefix']})")
                for u in progress(target_users, desc="Включение 2FA", total=len(target_users)):
                    uid = str(u.get("id"))
                    ok, err = client.set_user_2fa(uid, True)
                    if ok:
                        rep["users_2fa_enabled_ok"] += 1
                    else:
                        rep["users_2fa_enabled_fail"] += 1
                        log.error(f"User {uid} enable 2fa failed: {err}")
                print_report(rep)
            elif choice == "5":
                identifiers = prompt_identifiers()
                if not identifiers:
                    print("Шаг пропущен.")
                    continue
                rep = _new_report()
                users = client.list_users()
                rep["users_total"] = len(users)
                _filter_allowed(users, rep)
                _process_disable_identifiers(client, users, identifiers, rep)
                print_report(rep)
            elif choice == "0":
                print("Завершение работы.")
                return
            else:
                print("Неизвестный пункт меню.")
        except KeyboardInterrupt:
            print("\nПрервано пользователем.")
        except Exception as e:
            log.exception("Необработанная ошибка")
            print(f"Ошибка: {e}")

        # после завершения операции — пауза
        again = input("\nНажмите Enter, чтобы вернуться в меню, или 'q' для выхода: ").strip().lower()
        if again == "q":
            return


def main():
    print("=" * 60)
    print(" Yandex 360 — менеджер 2FA")
    print(" Лог-файл: " + log_file)
    print("=" * 60)
    try:
        token = ""
        while not token:
            token = masked_input("Введите OAuth-токен: ").strip()
            if not token:
                print("Токен не может быть пустым.")
        org_id = ""
        while not org_id:
            org_id = input("Введите ID организации: ").strip()
        main_menu(token, org_id)
    except KeyboardInterrupt:
        print("\nПрервано пользователем.")


if __name__ == "__main__":
    main()
