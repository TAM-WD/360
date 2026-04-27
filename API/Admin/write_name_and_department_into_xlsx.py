"""
Обогащает таблицу xlsx ФИО и подразделением сотрудника по email.

Для каждой строки из исходного файла берётся email (из колонки EMAIL_COLUMN),
ищется сотрудник через Yandex 360 API, и в результирующий файл
дописываются две колонки: "ФИО" (Фамилия Имя Отчество) и "Подразделение".

Инструкция по запуску:
1. Заполнить ORG_ID — id организации можно найти в панели управления admin.yandex.ru в левом нижнем углу.
2. Заполнить TOKEN — токен для доступа к API Yandex 360. Нужны права directory:read_users directory:read_departments
3. Заполнить FILEPATH — полный путь к исходному xlsx файлу.
4. Запустить скрипт.

Для установки необходимых библиотек нужно в терминале прописать команду 'pip install <lib>' (pip install requests, pip install openpyxl)
"""

import os
import re
import time
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# === Конфигурация ===
ORG_ID: str = ''
TOKEN: str = ''

# Путь к исходному xlsx файлу.
FILEPATH: str = r''

# Имя колонки в исходном файле, по значению которой ищем пользователя.
EMAIL_COLUMN: str = 'email'

# Имя листа в исходном файле; если None — берётся активный лист.
SHEET_NAME: Optional[str] = None

# Параметры API.
API_BASE: str = 'https://api360.yandex.net/directory/v1'
USERS_PER_PAGE: int = 1000
DEPARTMENTS_PER_PAGE: int = 1000
REQUEST_TIMEOUT_SEC: int = 30
SLEEP_BETWEEN_PAGES_SEC: float = 0.5


def build_session(token: str) -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=['GET'],
    )
    session.mount('https://', HTTPAdapter(max_retries=retry))
    session.headers.update({'Authorization': f'OAuth {token}'})
    return session


def fetch_paginated(
    session: requests.Session,
    url: str,
    items_key: str,
    per_page: int,
) -> list[dict]:
    """Универсальный обход постраничной выдачи Directory API (page/perPage)."""
    items: list[dict] = []
    page = 1
    while True:
        response = session.get(
            url,
            params={'page': page, 'perPage': per_page},
            timeout=REQUEST_TIMEOUT_SEC,
        )
        response.raise_for_status()
        data = response.json()
        chunk = data.get(items_key, [])
        items.extend(chunk)

        last_page = data.get('pages', 1)
        if page >= last_page or not chunk:
            break
        page += 1
        time.sleep(SLEEP_BETWEEN_PAGES_SEC)
    return items


def get_all_users(session: requests.Session, org_id: str) -> list[dict]:
    return fetch_paginated(
        session,
        url=f'{API_BASE}/org/{org_id}/users',
        items_key='users',
        per_page=USERS_PER_PAGE,
    )


def get_all_departments(session: requests.Session, org_id: str) -> list[dict]:
    return fetch_paginated(
        session,
        url=f'{API_BASE}/org/{org_id}/departments',
        items_key='departments',
        per_page=DEPARTMENTS_PER_PAGE,
    )


def build_users_index(users: list[dict]) -> dict[str, dict]:
    """Индекс пользователей по нижнему регистру email и всех алиасов."""
    index: dict[str, dict] = {}
    for user in users:
        emails: list[str] = []
        primary = user.get('email')
        if primary:
            emails.append(primary)
        emails.extend(user.get('aliases') or [])
        for contact in user.get('contacts') or []:
            if contact.get('type') == 'email' and contact.get('value'):
                emails.append(contact['value'])
        for email in emails:
            index[email.strip().lower()] = user
    return index


def build_departments_index(departments: list[dict]) -> dict[int, str]:
    return {dep['id']: dep.get('name', '') for dep in departments}

_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

def looks_like_email(value: object) -> bool:
    if not isinstance(value, str):
        return False
    return bool(_EMAIL_RE.match(value.strip()))


def is_data_row(email_cell: object) -> bool:
    """Строка считается строкой данных, если в колонке email пусто или валидный email.

    Любое другое непустое значение (например, "string", "тип", "—") трактуется
    как метаданные и пропускается — без привязки к конкретным значениям или
    номеру строки.
    """
    if email_cell is None:
        return True
    text = str(email_cell).strip()
    if not text:
        return True
    return looks_like_email(text)


def format_fio(user: dict) -> str:
    name = user.get('name') or {}
    parts = [name.get('last', ''), name.get('first', ''), name.get('middle', '')]
    return ' '.join(part for part in parts if part).strip()


def find_email_column_index(header_row: tuple, email_column: str) -> int:
    target = email_column.strip().lower()
    for idx, value in enumerate(header_row):
        if value is not None and str(value).strip().lower() == target:
            return idx
    raise ValueError(
        f'Не найдена колонка "{email_column}" в заголовке исходного файла. '
        f'Заголовок: {header_row}'
    )


def make_output_path(source_path: str) -> str:
    directory = os.path.dirname(os.path.abspath(source_path))
    base = os.path.splitext(os.path.basename(source_path))[0]
    timestamp = time.strftime('%Y-%m-%d_%H-%M-%S')
    return os.path.join(directory, f'{base}_with_fio_and_dep_{timestamp}.xlsx')


def enrich_workbook(
    src_path: str,
    sheet_name: Optional[str],
    email_column: str,
    users_by_email: dict[str, dict],
    departments_by_id: dict[int, str],
) -> str:
    src_wb = load_workbook(src_path, data_only=True, read_only=True)
    src_ws = src_wb[sheet_name] if sheet_name else src_wb.active

    src_ws.reset_dimensions()

    rows = src_ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise ValueError('Исходный файл пуст.')

    email_idx = find_email_column_index(header, email_column)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = src_ws.title

    out_header = list(header) + ['ФИО', 'Подразделение']
    out_ws.append(out_header)

    matched = 0
    total = 0
    skipped = 0
    for row in rows:
        email_cell = row[email_idx] if email_idx < len(row) else None

        if not is_data_row(email_cell):
            skipped += 1
            continue

        total += 1
        fio = ''
        department_name = ''

        if email_cell:
            user = users_by_email.get(str(email_cell).strip().lower())
            if user:
                matched += 1
                fio = format_fio(user)
                department_name = departments_by_id.get(user.get('departmentId'), '')

        out_ws.append(list(row) + [fio, department_name])

    out_path = make_output_path(src_path)
    out_wb.save(out_path)
    src_wb.close()

    print(
        f'Обработано строк: {total}, найдено сотрудников: {matched}, '
        f'пропущено не-данных: {skipped}'
    )
    return out_path


def main() -> None:
    if not ORG_ID or not TOKEN:
        raise SystemExit('Заполните ORG_ID и TOKEN в начале скрипта.')
    if not FILEPATH:
        raise SystemExit('Заполните FILEPATH — путь к исходному xlsx файлу.')
    if not os.path.isfile(FILEPATH):
        raise SystemExit(f'Файл не найден: {FILEPATH}')

    session = build_session(TOKEN)

    print('Получаю список сотрудников...')
    users = get_all_users(session, ORG_ID)
    print(f'  получено: {len(users)}')

    print('Получаю список подразделений...')
    departments = get_all_departments(session, ORG_ID)
    print(f'  получено: {len(departments)}')

    users_by_email = build_users_index(users)
    departments_by_id = build_departments_index(departments)

    out_path = enrich_workbook(
        src_path=FILEPATH,
        sheet_name=SHEET_NAME,
        email_column=EMAIL_COLUMN,
        users_by_email=users_by_email,
        departments_by_id=departments_by_id,
    )
    print(f'Готово. Результат: {out_path}')


if __name__ == '__main__':
    main()
