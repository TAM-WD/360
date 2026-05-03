### Скрипт v2 — выгрузка пользователей организации Яндекс 360
# Используется API Яндекс 360 для бизнеса
# Поддержка: departments API (все параметры), CSV/XLSX, интерактивное меню, отчёт

import os
import sys
import csv
import getpass
import time
import requests
from requests.adapters import HTTPAdapter, Retry
from datetime import datetime

# ─── Константы ────────────────────────────────────────────────────────────────
PERPAGE = 1000
BASE_URL = "https://api360.yandex.net/directory/v1/org"

# ─── Глобальные настройки (заполняются через меню) ────────────────────────────
TOKEN = ""
ORGID = ""
OUTPUT_FORMAT = "csv"  # "csv" или "xlsx"


# ═══════════════════════════════════════════════════════════════════════════════
#  API-функции
# ═══════════════════════════════════════════════════════════════════════════════

def _session():
    """Создаёт requests.Session с retry-политикой."""
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    s = requests.Session()
    s.mount("https://", HTTPAdapter(max_retries=retries))
    return s


def _headers():
    return {"Authorization": f"OAuth {TOKEN}"}


# ── Пользователи ─────────────────────────────────────────────────────────────

def api360_get_users(orgid, page, per_page=PERPAGE):
    """GET /directory/v1/org/{orgId}/users — список сотрудников."""
    url = f"{BASE_URL}/{orgid}/users"
    params = {"page": page, "perPage": per_page}
    r = _session().get(url, params=params, headers=_headers())
    r.raise_for_status()
    data = r.json()
    return data.get("users", []), data.get("pages", 1)


# ── Подразделения ─────────────────────────────────────────────────────────────
# Справка: https://yandex.ru/dev/api360/doc/ru/directory/get-departments
#
# Query-параметры:
#   page      — номер страницы (начиная с 1)
#   perPage   — кол-во записей на странице (макс. 1000)
#   parentId  — ID родительского подразделения (фильтр)
#   orderBy   — поле сортировки: "id" или "name"

def api360_get_departments(orgid, page=1, per_page=PERPAGE,
                           parent_id=None, order_by=None):
    """GET /directory/v1/org/{orgId}/departments — список подразделений.

    Поддерживает ВСЕ query-параметры из документации:
      • page, perPage — пагинация
      • parentId      — фильтр по родительскому подразделению
      • orderBy       — сортировка ("id" | "name")
    """
    url = f"{BASE_URL}/{orgid}/departments"
    params = {"page": page, "perPage": per_page}
    if parent_id is not None:
        params["parentId"] = parent_id
    if order_by is not None:
        params["orderBy"] = order_by
    r = _session().get(url, params=params, headers=_headers())
    r.raise_for_status()
    data = r.json()
    return data.get("departments", []), data.get("pages", 1)


def api360_get_department_info(orgid, department_id):
    """GET /directory/v1/org/{orgId}/departments/{depId} — информация о подразделении."""
    url = f"{BASE_URL}/{orgid}/departments/{department_id}"
    r = _session().get(url, headers=_headers())
    r.raise_for_status()
    return r.json()


def fetch_all_departments(orgid, parent_id=None, order_by=None):
    """Загружает ВСЕ подразделения (с пагинацией) и возвращает dict {id: dept}."""
    departments = {}
    page = 1
    total_pages = 1
    while page <= total_pages:
        deps, total_pages = api360_get_departments(
            orgid, page=page, per_page=PERPAGE,
            parent_id=parent_id, order_by=order_by,
        )
        for d in deps:
            departments[d["id"]] = d
        page += 1
    return departments


# ═══════════════════════════════════════════════════════════════════════════════
#  Утилиты вывода
# ═══════════════════════════════════════════════════════════════════════════════

def mask_token(token: str) -> str:
    """Маскирует токен символами для отображения."""
    if not token:
        return "(пусто)"
    symbols = "█▓▒░◆◇●○■□▲△"
    masked = ""
    for i, ch in enumerate(token):
        masked += symbols[i % len(symbols)]
    return masked


def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def print_header():
    print("=" * 60)
    print("  Яндекс 360 — Выгрузка пользователей организации  v2.0")
    print("=" * 60)


def print_current_settings():
    print()
    print("  Текущие настройки:")
    print(f"    Токен  : {mask_token(TOKEN)}")
    print(f"    Org ID : {ORGID if ORGID else '(не задан)'}")
    print(f"    Формат : {OUTPUT_FORMAT.upper()}")
    print()


# ═══════════════════════════════════════════════════════════════════════════════
#  Интерактивное меню
# ═══════════════════════════════════════════════════════════════════════════════

def menu_set_token():
    global TOKEN
    print()
    print("  Введите OAuth-токен (ввод скрыт):")
    token = getpass.getpass("  Токен > ")
    if token.strip():
        TOKEN = token.strip()
        print(f"  ✔ Токен сохранён: {mask_token(TOKEN)}")
    else:
        print("  ✘ Токен не изменён (пустой ввод).")
    input("\n  Нажмите Enter для продолжения...")


def menu_set_orgid():
    global ORGID
    print()
    orgid = input("  Введите ID организации (orgId) > ").strip()
    if orgid:
        ORGID = orgid
        print(f"  ✔ Org ID сохранён: {ORGID}")
    else:
        print("  ✘ Org ID не изменён (пустой ввод).")
    input("\n  Нажмите Enter для продолжения...")


def menu_set_format():
    global OUTPUT_FORMAT
    print()
    print("  Выберите формат выгрузки:")
    print("    1. CSV  (.csv)")
    print("    2. XLSX (.xlsx)")
    print("    0. Назад")
    choice = input("  > ").strip()
    if choice == "1":
        OUTPUT_FORMAT = "csv"
        print("  ✔ Формат: CSV")
    elif choice == "2":
        OUTPUT_FORMAT = "xlsx"
        print("  ✔ Формат: XLSX")
    elif choice == "0":
        return
    else:
        print("  ✘ Неверный выбор.")
    input("\n  Нажмите Enter для продолжения...")


def menu_run_export():
    if not TOKEN:
        print("\n  ✘ Сначала укажите токен (пункт 1).")
        input("\n  Нажмите Enter для продолжения...")
        return
    if not ORGID:
        print("\n  ✘ Сначала укажите Org ID (пункт 2).")
        input("\n  Нажмите Enter для продолжения...")
        return
    run_export()


def main_menu():
    while True:
        clear_screen()
        print_header()
        print_current_settings()
        print("  ┌─────────────────────────────────────┐")
        print("  │  1. Ввести токен                     │")
        print("  │  2. Ввести Org ID                    │")
        print("  │  3. Выбрать формат (CSV / XLSX)      │")
        print("  │  4. ▶ Запустить выгрузку              │")
        print("  │  0. Выход                            │")
        print("  └─────────────────────────────────────┘")
        choice = input("  Выберите пункт > ").strip()
        if choice == "1":
            menu_set_token()
        elif choice == "2":
            menu_set_orgid()
        elif choice == "3":
            menu_set_format()
        elif choice == "4":
            menu_run_export()
        elif choice == "0":
            print("\n  До свидания!")
            sys.exit(0)
        else:
            print("  ✘ Неверный выбор.")
            time.sleep(1)


# ═══════════════════════════════════════════════════════════════════════════════
#  Основная логика выгрузки
# ═══════════════════════════════════════════════════════════════════════════════

# Порядок колонок (первые 7 — по ТЗ, остальные — дополнительные)
FIELD_NAMES = [
    "id",
    "nickname",
    "email",
    "firstName",
    "lastName",
    "departmentName",
    "departmentId",
    # --- дополнительные ---
    "middleName",
    "position",
    "isEnabled",
    "isEnabledUpdatedAt",
    "isAdmin",
    "gender",
    "about",
    "birthday",
    "aliases",
    "groups",
    "externalId",
    "isRobot",
    "isDismissed",
    "timezone",
    "language",
    "createdAt",
    "updatedAt",
    "contacts",
]


def flatten_user(user: dict, dept_map: dict) -> dict:
    """Преобразует объект пользователя API в плоский dict для записи."""
    name = user.get("name", {})
    dep_id = user.get("departmentId", "")
    dep_name = ""
    if dep_id and dep_id in dept_map:
        dep_name = dept_map[dep_id].get("name", "")

    flat = {
        "id": user.get("id", ""),
        "nickname": user.get("nickname", ""),
        "email": user.get("email", ""),
        "firstName": name.get("first", ""),
        "lastName": name.get("last", ""),
        "departmentName": dep_name,
        "departmentId": dep_id,
        "middleName": name.get("middle", ""),
        "position": user.get("position", ""),
        "isEnabled": user.get("isEnabled", ""),
        "isEnabledUpdatedAt": user.get("isEnabledUpdatedAt", ""),
        "isAdmin": user.get("isAdmin", ""),
        "gender": user.get("gender", ""),
        "about": user.get("about", ""),
        "birthday": user.get("birthday", ""),
        "aliases": "; ".join(user.get("aliases", [])) if isinstance(user.get("aliases"), list) else user.get("aliases", ""),
        "groups": "; ".join(str(g.get("id", g) if isinstance(g, dict) else g) for g in user.get("groups", [])) if isinstance(user.get("groups"), list) else user.get("groups", ""),
        "externalId": user.get("externalId", ""),
        "isRobot": user.get("isRobot", ""),
        "isDismissed": user.get("isDismissed", ""),
        "timezone": user.get("timezone", ""),
        "language": user.get("language", ""),
        "createdAt": user.get("createdAt", ""),
        "updatedAt": user.get("updatedAt", ""),
        "contacts": "; ".join(
            f"{c.get('type', '')}:{c.get('value', '')}"
            for c in user.get("contacts", [])
        ) if isinstance(user.get("contacts"), list) else user.get("contacts", ""),
    }
    return flat


def write_csv(file_path: str, rows: list[dict]):
    """Записывает данные в CSV (UTF-8 BOM, разделитель ;)."""
    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, FIELD_NAMES, extrasaction="ignore", delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(file_path: str, rows: list[dict]):
    """Записывает данные в XLSX с помощью openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ✘ Библиотека openpyxl не установлена.")
        print("    Установите: pip install openpyxl")
        input("\n  Нажмите Enter для продолжения...")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    # Заголовки
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, field in enumerate(FIELD_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(FIELD_NAMES, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))

    # Автоширина колонок
    for col_idx, field in enumerate(FIELD_NAMES, 1):
        max_len = len(str(field))
        for row_idx in range(2, min(len(rows) + 2, 102)):  # проверяем первые 100 строк
            val = str(rows[row_idx - 2].get(field, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    wb.save(file_path)


def run_export():
    """Основная процедура выгрузки."""
    clear_screen()
    print_header()
    print()
    print("  ▶ Запуск выгрузки...")
    print()

    start_time_dt = datetime.now()
    start_time = start_time_dt.strftime("%Y-%m-%d_%H-%M-%S")
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # ── 1. Загрузка подразделений ─────────────────────────────────────────────
    print("  [1/3] Загрузка подразделений...")
    try:
        dept_map = fetch_all_departments(ORGID, order_by="name")
        dept_count = len(dept_map)
        print(f"         Найдено подразделений: {dept_count}")
    except requests.exceptions.HTTPError as e:
        print(f"  ✘ Ошибка при загрузке подразделений: {e}")
        print("    Возможно, у токена нет прав directory:read_departments.")
        print("    Продолжаю без названий подразделений...")
        dept_map = {}
        dept_count = 0
    except Exception as e:
        print(f"  ✘ Ошибка: {e}")
        dept_map = {}
        dept_count = 0

    # ── 2. Загрузка пользователей ─────────────────────────────────────────────
    print("  [2/3] Загрузка пользователей...")
    all_rows = []
    page = 1
    total_pages = 1
    user_count = 0

    try:
        while page <= total_pages:
            users, total_pages = api360_get_users(ORGID, page)
            for user in users:
                flat = flatten_user(user, dept_map)
                all_rows.append(flat)
                user_count += 1
            print(f"         Страница {page}/{total_pages} — загружено {len(users)} пользователей")
            page += 1
    except requests.exceptions.HTTPError as e:
        print(f"  ✘ Ошибка API: {e}")
        if not all_rows:
            input("\n  Нажмите Enter для возврата в меню...")
            return
        print("    Сохраняю то, что удалось загрузить...")
    except Exception as e:
        print(f"  ✘ Ошибка: {e}")
        if not all_rows:
            input("\n  Нажмите Enter для возврата в меню...")
            return

    # ── 3. Сохранение файла ───────────────────────────────────────────────────
    ext = OUTPUT_FORMAT
    file_name = f"users_orgid_{ORGID}_{start_time}.{ext}"
    file_path = os.path.join(script_dir, file_name)

    print(f"  [3/3] Сохранение в {ext.upper()}...")

    if OUTPUT_FORMAT == "csv":
        write_csv(file_path, all_rows)
    else:
        write_xlsx(file_path, all_rows)

    end_time_dt = datetime.now()
    elapsed = end_time_dt - start_time_dt

    # ── Отчёт ─────────────────────────────────────────────────────────────────
    print()
    print("  ╔══════════════════════════════════════════════════════╗")
    print("  ║                    ОТЧЁТ ВЫГРУЗКИ                   ║")
    print("  ╠══════════════════════════════════════════════════════╣")
    print(f"  ║  Сотрудников выгружено  : {user_count:<26}║")
    print(f"  ║  Подразделений найдено  : {dept_count:<26}║")
    print(f"  ║  Формат файла           : {ext.upper():<26}║")
    print(f"  ║  Файл                   : {file_name:<26}║")
    print(f"  ║  Время выполнения       : {str(elapsed).split('.')[0]:<26}║")
    print("  ╚══════════════════════════════════════════════════════╝")
    print()
    print(f"  Файл сохранён: {file_path}")
    print()

    # Возврат в начало
    while True:
        print("  ┌─────────────────────────────────────┐")
        print("  │  1. Вернуться в главное меню         │")
        print("  │  0. Выход                            │")
        print("  └─────────────────────────────────────┘")
        choice = input("  > ").strip()
        if choice == "1":
            return  # возврат в main_menu
        elif choice == "0":
            print("\n  До свидания!")
            sys.exit(0)
        else:
            print("  ✘ Неверный выбор.")


# ═══════════════════════════════════════════════════════════════════════════════
#  Точка входа
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    try:
        main_menu()
    except KeyboardInterrupt:
        print("\n\n  Прервано пользователем (Ctrl+C).")
        sys.exit(0)
