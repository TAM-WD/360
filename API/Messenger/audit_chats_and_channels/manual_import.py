from __future__ import annotations

import csv
import glob
import hashlib
import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone

from text_utils import (clean_text, looks_like_email,
                        looks_like_messenger_bot_login, position_quality_flags)

log = logging.getLogger("manual")

# В Excel дата хранится числом дней. База 1899-12-30 — с поправкой на
# известную ошибку Excel с 1900 годом.
EXCEL_EPOCH = datetime(1899, 12, 30)

TYPE_MAP = {
    "канал": "channel", "channel": "channel",
    "групповой чат": "group", "групповой": "group", "группа": "group",
    "group": "group",
    "приватный": "private", "личный чат": "private", "private": "private",
}

HEADER_ALIASES = {
    "messenger_chat.created": "date_raw", "дата": "date_raw", "date": "date_raw",
    "chat_name": "chat_name", "название чата": "chat_name",
    "chat_description": "chat_description", "описание": "chat_description",
    "тип": "chat_type", "type": "chat_type",
    "full name": "full_name", "фио": "full_name",
    "job position": "position", "должность": "position",
    "email": "identity", "почта": "identity", "логин": "identity",
    "login": "identity",
    "role": "role", "роль": "role",
    "added by": "added_by", "кем добавлен": "added_by",
}


# =========================================================================
# поиск файла
# =========================================================================
def resolve_manual_path(path: str) -> str:
    """Ищет файл, учитывая ~, регистр и то, что macOS хранит имена файлов
    в другой форме записи Unicode. При неудаче объясняет, что делать."""
    expanded = os.path.expanduser(os.path.expandvars(path))
    if os.path.isfile(expanded):
        return expanded

    directory = os.path.dirname(expanded) or "."
    target = os.path.basename(expanded)

    if os.path.isdir(directory):
        for form in ("NFC", "NFD"):
            candidate = os.path.join(directory, unicodedata.normalize(form, target))
            if os.path.isfile(candidate):
                return candidate
        target_folded = unicodedata.normalize("NFC", target).casefold()
        try:
            for entry in os.listdir(directory):
                if unicodedata.normalize("NFC", entry).casefold() == target_folded:
                    return os.path.join(directory, entry)
        except OSError:
            pass

    nearby: list[str] = []
    for search_dir in (os.getcwd(), os.path.expanduser("~/Downloads")):
        for pattern in ("*.xlsx", "*.xlsm", "*.csv"):
            nearby.extend(glob.glob(os.path.join(search_dir, pattern)))

    hint = ""
    if nearby:
        listing = "\n".join(f"    {item}" for item in sorted(set(nearby))[:15])
        hint = f"\n\nРядом нашлись такие таблицы:\n{listing}"

    raise FileNotFoundError(
        f"Не удалось найти файл: {path!r}\n"
        f"Искали относительно каталога: {os.getcwd()}"
        f"{hint}\n\n"
        f"Укажите полный путь либо положите файл рядом со скриптом:\n"
        f"    cp ~/Downloads/<файл>.xlsx ./manual.xlsx\n"
        f"    python cli.py validate --manual manual.xlsx --manual-tz \"+03:00\""
    )


# =========================================================================
# модели
# =========================================================================
@dataclass
class ManualRow:
    row_no: int
    date: datetime | None
    chat_name: str
    chat_description: str | None
    chat_type: str | None
    full_name: str | None
    position: str | None
    identity: str | None                # адрес почты или логин
    identity_kind: str = "unknown"      # email | bot_login | login | unknown
    is_bot: bool = False                # догадка по виду логина
    role: str | None = None
    added_by: str | None = None
    quality_flags: list[str] = field(default_factory=list)


@dataclass
class ManualChat:
    chat_name: str
    chat_type: str | None
    description: str | None
    manual_key: str
    created_at: datetime | None = None
    members: dict[str, ManualRow] = field(default_factory=dict)

    @property
    def bots_count(self) -> int:
        return sum(1 for row in self.members.values() if row.is_bot)


@dataclass
class ManualImportResult:
    chats: dict[str, ManualChat] = field(default_factory=dict)
    rows_total: int = 0
    rows_ok: int = 0
    errors: list[tuple[int, str]] = field(default_factory=list)
    conflicts: list[dict] = field(default_factory=list)
    quality_issues: list[dict] = field(default_factory=list)
    source_file: str | None = None

    @property
    def bots_total(self) -> int:
        return sum(chat.bots_count for chat in self.chats.values())


# =========================================================================
# даты
# =========================================================================
def _parse_tz(offset: str) -> timezone:
    match = re.fullmatch(r"([+-])(\d{2}):(\d{2})", (offset or "").strip())
    if not match:
        return timezone.utc
    sign = 1 if match.group(1) == "+" else -1
    return timezone(sign * timedelta(hours=int(match.group(2)),
                                     minutes=int(match.group(3))))


def excel_serial_to_dt(value, tz_offset: str) -> datetime | None:
    """Число из Excel, datetime или строка ISO -> дата с часовым поясом."""
    if value is None or value == "":
        return None
    tzinfo = _parse_tz(tz_offset)

    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=tzinfo)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        serial = float(value)
        if not (1 < serial < 200000):
            raise ValueError(f"число {serial} не похоже на дату Excel")
        return (EXCEL_EPOCH + timedelta(days=serial)).replace(tzinfo=tzinfo)

    text = str(value).strip()
    try:
        return excel_serial_to_dt(float(text.replace(",", ".")), tz_offset)
    except ValueError:
        pass
    try:
        parsed = datetime.fromisoformat(text)
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=tzinfo)
    except ValueError:
        raise ValueError(f"не удалось прочитать дату: {value!r}") from None


# =========================================================================
# нормализация
# =========================================================================
def normalize_name(name: str | None) -> str:
    return (clean_text(name) or "").casefold()


def normalize_identity(value: str | None) -> str | None:
    cleaned = clean_text(value)
    return cleaned.casefold() if cleaned else None


def classify_identity(identity: str | None) -> tuple[str, bool]:
    """Определяет вид идентификатора -> (вид, похож_на_бота)."""
    if not identity:
        return "unknown", False
    if looks_like_messenger_bot_login(identity):
        return "bot_login", True
    if looks_like_email(identity):
        return "email", False
    return "login", False


def make_manual_key(chat_name: str, chat_type: str | None) -> str:
    raw = f"{normalize_name(chat_name)}|{chat_type or ''}"
    return "manual::" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


# =========================================================================
# чтение файла
# =========================================================================
def _map_headers(header_row: list) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, cell in enumerate(header_row):
        key = (clean_text(cell) or "").casefold()
        if key in HEADER_ALIASES:
            mapping[index] = HEADER_ALIASES[key]
    return mapping


def _iter_raw_rows(path: str, sheet: str | None):
    extension = os.path.splitext(path)[1].lower()

    if extension in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet:
            if sheet not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                workbook.close()
                raise ValueError(
                    f"В файле нет листа {sheet!r}. Есть такие листы: {available}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        for row in worksheet.iter_rows(values_only=True):
            yield list(row)
        workbook.close()

    elif extension == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            except csv.Error:
                dialect = csv.excel
            for row in csv.reader(handle, dialect):
                yield row
    else:
        raise ValueError(
            f"С таким форматом файла работать не умеем: {extension}. "
            f"Подойдёт .xlsx, .xlsm или .csv")


# =========================================================================
# основной разбор
# =========================================================================
def load_manual(path: str, *, sheet: str | None = None, tz_offset: str = "+00:00",
                date_semantics: str = "member_added") -> ManualImportResult:
    resolved = resolve_manual_path(path)
    result = ManualImportResult(source_file=resolved)
    header_seen = False
    col_map: dict[int, str] = {}

    for raw in _iter_raw_rows(resolved, sheet):
        if not header_seen:
            col_map = _map_headers(raw)
            if "chat_name" not in col_map.values():
                recognized = sorted(set(col_map.values()))
                raise ValueError(
                    "В таблице не нашлась колонка с названием чата (chat_name). "
                    f"Удалось распознать такие колонки: {recognized}. "
                    f"Первая строка файла: {[clean_text(c) for c in raw][:10]}")
            header_seen = True
            continue

        if all(cell is None or str(cell).strip() == "" for cell in raw):
            continue

        result.rows_total += 1
        row_no = result.rows_total + 1          # +1 на строку заголовка

        values = {name: (raw[index] if index < len(raw) else None)
                  for index, name in col_map.items()}

        chat_name = clean_text(values.get("chat_name"))
        if not chat_name:
            result.errors.append((row_no, "не указано название чата"))
            continue

        try:
            row_date = excel_serial_to_dt(values.get("date_raw"), tz_offset)
        except ValueError as exc:
            result.errors.append((row_no, str(exc)))
            row_date = None

        raw_type = clean_text(values.get("chat_type"))
        chat_type = TYPE_MAP.get((raw_type or "").casefold())
        if raw_type and chat_type is None:
            result.errors.append(
                (row_no, f"незнакомый тип чата: {raw_type!r}"))

        identity = normalize_identity(values.get("identity"))
        if not identity:
            result.errors.append(
                (row_no, "не указан адрес почты или логин — участника не опознать"))

        identity_kind, is_bot = classify_identity(identity)
        full_name = clean_text(values.get("full_name"))
        position = clean_text(values.get("position"))

        # Замечания к качеству. Данные не исправляем, только помечаем.
        # К ботам проверки не применяем: ФИО и должности у них не бывает,
        # а логин вместо адреса почты для бота — норма.
        flags: list[str] = []
        if not is_bot:
            flags += position_quality_flags(position)
            if identity_kind == "login":
                flags.append("identity_is_login_not_email")
            if identity_kind == "unknown" and identity:
                flags.append("identity_unrecognized")

        for flag in flags:
            result.quality_issues.append({
                "row_no": row_no,
                "chat_name": chat_name,
                "identity": identity or "",
                "flag": flag,
                "value": position or full_name or "",
            })

        row = ManualRow(
            row_no=row_no, date=row_date, chat_name=chat_name,
            chat_description=clean_text(values.get("chat_description")),
            chat_type=chat_type, full_name=full_name, position=position,
            identity=identity, identity_kind=identity_kind, is_bot=is_bot,
            role=clean_text(values.get("role")),
            added_by=normalize_identity(values.get("added_by")),
            quality_flags=flags,
        )

        manual_key = make_manual_key(chat_name, chat_type)
        chat = result.chats.setdefault(manual_key, ManualChat(
            chat_name=chat_name, chat_type=chat_type,
            description=row.chat_description, manual_key=manual_key))

        if row.chat_description and chat.description and \
                row.chat_description != chat.description:
            result.conflicts.append({
                "kind": "description_mismatch", "chat_name": chat_name,
                "row_no": row_no,
                "detail": f"{chat.description!r} и {row.chat_description!r}"})
        if not chat.description and row.chat_description:
            chat.description = row.chat_description

        if identity:
            previous = chat.members.get(identity)
            if previous is None:
                chat.members[identity] = row
            else:
                result.conflicts.append({
                    "kind": "duplicate_member", "chat_name": chat_name,
                    "row_no": row_no,
                    "detail": f"{identity} встречается в строках "
                              f"{previous.row_no} и {row_no}"})
                if row.date and (previous.date is None or row.date < previous.date):
                    chat.members[identity] = row

        result.rows_ok += 1

    if not header_seen:
        raise ValueError(f"Файл пустой или в нём нет ни одной строки: {resolved}")

    for chat in result.chats.values():
        dates = [row.date for row in chat.members.values() if row.date]
        if not dates:
            continue
        chat.created_at = min(dates)
        if date_semantics == "chat_created" and len(set(dates)) > 1:
            result.conflicts.append({
                "kind": "chat_created_ambiguous", "chat_name": chat.chat_name,
                "row_no": "",
                "detail": f"у участников {len(set(dates))} разных дат, "
                          f"за дату создания взяли самую раннюю — "
                          f"{chat.created_at.strftime('%d.%m.%Y')}"})

    log.info("Таблица прочитана: строк %s, приняли %s, чатов %s, ботов %s, "
             "ошибок %s, противоречий %s, замечаний %s",
             result.rows_total, result.rows_ok, len(result.chats),
             result.bots_total, len(result.errors), len(result.conflicts),
             len(result.quality_issues))
    return result
